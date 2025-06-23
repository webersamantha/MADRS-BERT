"""
Author: Samantha Weber (samantha.weber@bli.uzh.ch)

Description: This project has been generated within the MULTICAST [https://www.multicast.uzh.ch/en.html] project at University of Zurich and University Hospital of Psychiatry Zurich.

Main Steps in the Workflow
	1.	Extracts Audio from Videos
	    •	Identifies subject folders (MC_1XXXX).
	    •	Searches for video files ending with "_MADRS.mp4".
	    •	Extracts the audio track and saves it in WAV format.
	    •	Converts the audio to mono-channel if needed.
	2.	Performs Speaker Diarization
	    •	Uses PyAnnote’s speaker diarization pipeline to identify different speakers.
	    •	Segments the audio file into speaker-specific parts.
	3.	Transcribes Speech to Text
	    •	Uses Whisper (OpenAI) to transcribe each diarized audio segment.
	    •	Saves the transcriptions along with subject IDs and speakers in an Excel file.
	4.	Tracks Processed Subjects in an Excel File
	    •	Checks if a subject is already processed (avoiding duplicate work).
	    •	Saves processed subject information into an Excel log (01_all_transcriptions.xlsx).

Clinical interviews were conducted with patients (MADRS), which were recorded on video. Language: SwissGerman and German
"""

# Import necessary libraries
from pyannote.audio import Pipeline
from moviepy.editor import VideoFileClip
from pydub import AudioSegment
from openpyxl import load_workbook, Workbook
import os
import certifi
import whisper
from huggingface_hub import login
import logging
from tqdm import tqdm  # For progress tracking

# Configure logging
logging.basicConfig(level=logging.INFO)

# Path to the token file
token_file_path = "YOUR_HF_TOKEN_HERE"

# Read the token from the file
with open(token_file_path, 'r') as file:
    login_token = file.read().strip()

# Use the token for login
login(token=login_token)

# Ensure the correct certificate store is used
os.environ['SSL_CERT_FILE'] = certifi.where()

# Paths
video_path='./Video'
audio_path = "./01_Raw"
audio_path_filtered = "./02_Filtered"
xlsx_file_path = "./03_Text_Extracted/01_all_transcriptions.xlsx"

# Function to check if a subject exists in the Excel file
def subject_exists_in_excel(subject_code, xlsx_file_path):
    if os.path.exists(xlsx_file_path):
        workbook = load_workbook(xlsx_file_path)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] == subject_code:
                return True
    return False

# Function to save transcription to a common Excel file
def save_transcription_to_excel(subject_code, speaker, transcription, xlsx_file_path):
    if os.path.exists(xlsx_file_path):
        # Load the existing workbook
        workbook = load_workbook(xlsx_file_path)
        sheet = workbook.active
        next_row = sheet.max_row + 1
    else:
        # Create a new workbook and add headers
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Transcription"
        # Write headers
        sheet["A1"] = "Subject Code"
        sheet["B1"] = "Speaker"
        sheet["C1"] = "Transcription"
        next_row = 2
    # Write data
    sheet[f"A{next_row}"] = subject_code
    sheet[f"B{next_row}"] = speaker
    sheet[f"C{next_row}"] = transcription

    # Save the workbook
    workbook.save(xlsx_file_path)
    print(f"Transcription saved to {xlsx_file_path}")

# Process each subject folder
subject_folders = [f for f in os.listdir(video_path) if f.startswith("MC_1")]

for subject_folder in subject_folders:
    subject_path = os.path.join(video_path, subject_folder)
    # Look for the video file
    for file in os.listdir(subject_path):
        if file.endswith("_MADRS.mp4"):
            video_file = os.path.join(subject_path, file)
            audio_file = os.path.join(audio_path, file.replace(".mp4", ".wav"))
            filtered_audio_file = os.path.join(audio_path_filtered, "f_" + file.replace(".mp4", ".wav"))
            
            # Extract full subject code (e.g., MC_1XXX)
            subject_code = "_".join(os.path.basename(filtered_audio_file).split("_")[1:4]) 

            # Check if the subject code already exists in the Excel file
            if subject_exists_in_excel(subject_code, xlsx_file_path):
                print(f"Subject {subject_code} already processed. Skipping...")
                continue

            # 1. Conversion Video to Audio ------------------------------------------------------------------------------
            # Check if the audio file already exists
            if not os.path.exists(audio_file):
                print(f"Extracting audio from {video_file}...")
                try:
                    # Extract Audiofile
                    video = VideoFileClip(video_file)
                    audio = video.audio
                    # Write audiofiles
                    audio.write_audiofile(audio_file)
                except Exception as e:
                    logging.error(f"Error extracting audio from video: {e}")
                    raise

            if not os.path.exists(filtered_audio_file):
                try:
                    audio = AudioSegment.from_wav(audio_file)

                    # Reduce to one channel if necessary
                    audio = audio.set_channels(1)
                    audio.export(filtered_audio_file, format="wav")
                except Exception as e:
                    logging.error(f"Error processing audio file: {e}")
                    raise

            # 2. Conversion Speech to Text ------------------------------------------------------------------------------
            
            logging.info(f"Starting diarization for subject {subject_code}...")
            
            # Diarization
            pipeline = Pipeline.from_pretrained("pyannote/speaker-diarization-3.1", use_auth_token=login_token)

            # Apply diarization
            diarization = pipeline(filtered_audio_file)
            
            logging.info(f"Diarization completed for subject {subject_code}.")

            # Transcription
            # Load the pre-trained model
            model_name = "large"
            model = whisper.load_model(model_name)

            # Process diarization and transcribe each segment
            original_audio = AudioSegment.from_wav(filtered_audio_file)
            segments = list(diarization.itertracks(yield_label=True))
            for segment, _, speaker in tqdm(segments, desc=f"Processing {subject_code}", leave=True):
                start_ms = int(segment.start * 1000)
                end_ms = int(segment.end * 1000)
                segment_audio = original_audio[start_ms:end_ms]
                # Save diarized segment with correct naming
                segment_audio_path = os.path.join(audio_path_filtered, f"{subject_code}_{speaker}_{start_ms}_{end_ms}.wav")
                segment_audio.export(segment_audio_path, format="wav")
                
                # Transcribe the audio segment
                result = model.transcribe(segment_audio_path, language="de")
                transcription = result["text"]
                print(f"Speaker {speaker} Transcription:", transcription)

                # Save the transcription to the common Excel file
                save_transcription_to_excel(subject_code, speaker, transcription, xlsx_file_path)

print(f"Processing completed. Transcriptions saved to {xlsx_file_path}")