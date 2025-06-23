"""
Author: Samantha Weber (samantha.weber@bli.uzh.ch)

Description: This script is designed to process video recordings of clinical interviews conducted with patients using the Montgomery-Åsberg Depression Rating Scale (MADRS). The interviews, recorded in Swiss German or German, are stored in the ./Video folder. 
The script performs the following key tasks:
	1.	Video Processing & Audio Extraction
	    •	Identifies video files ending with "_MADRS.mp4" in participant folders.
	    •	Extracts the audio track from each video and saves it as a .wav file in the 01_Raw directory.
	2.	Audio Preprocessing
	    •	Converts extracted audio to mono-channel format for consistency.
	    •	Saves the filtered version in the 02_Filtered directory.
	3.	Subject Tracking & Excel Logging
	    •	Checks if a participant has already been processed by referencing an Excel file (01_ManualChecks.xlsx).
	    •	Prevents duplicate processing by skipping previously logged subjects.
	    •	Updates the Excel file after successful audio extraction.
	4.	Error Handling & Logging
	    •	Implements robust error handling for video/audio processing failures.
	    •	Uses logging to track processing status and potential issues.
"""

# Import necessary libraries
from moviepy.editor import VideoFileClip
from pydub import AudioSegment
from openpyxl import load_workbook
import os
import certifi
import logging
from pathlib import Path


# Configure logging
logging.basicConfig(level=logging.INFO)

# Ensure the correct certificate store is used
os.environ['SSL_CERT_FILE'] = certifi.where()

# Paths
video_path = Path("./Videos")
audio_path = Path("./Audios/01_Raw")
audio_path_filtered = Path("./Audios/02_Filtered")
xlsx_file_path = Path("./01_ManualChecks.xlsx")

# Function to check if a subject exists in the Excel file
def subject_exists_in_excel(subject_code, xlsx_file_path):
    if os.path.exists(xlsx_file_path):
        workbook = load_workbook(xlsx_file_path)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] == subject_code:
                return True
    return False

# Function to add the subject code to the Excel file
def add_subject_to_excel(subject_code, xlsx_file_path):
    workbook = load_workbook(xlsx_file_path)
    sheet = workbook.active
    next_row = sheet.max_row + 1
    sheet.cell(row=next_row, column=1, value=subject_code)
    workbook.save(xlsx_file_path)
    print(f"Added subject code {subject_code} to Excel file.")

# Process each subject folder in the T0 path
subject_folders = [f for f in os.listdir(video_path) if f.startswith("MC_1")] #MC_1 is the start of the participant code. 

for subject_folder in subject_folders:
    subject_path = os.path.join(video_path, subject_folder)
    # Look for the video file
    for file in os.listdir(subject_path):
        if file.endswith("_MADRS.mp4"):
            video_file = os.path.join(subject_path, file)
            
            # Extract full subject code (e.g., MC_XXXX)
            subject_code_parts = os.path.basename(file).split("_")
            subject_code = "_".join(subject_code_parts[0:2])  
            audio_file_name = f"{subject_code}_MADRS.wav"
            
            audio_file = os.path.join(audio_path, audio_file_name)
            filtered_audio_file = os.path.join(audio_path_filtered, "f_" + audio_file_name)

            # Check if the subject code already exists in the Excel file
            if subject_exists_in_excel(subject_code, xlsx_file_path):
                print(f"Subject {subject_code} already processed. Skipping...")
                continue

            # 1. Conversion Video to Audio ------------------------------------------------------------------------------
            # Check if the audio file already exists
            if not os.path.exists(audio_file):
                print(f"Extracting audio from {video_file}...")
                try:
                    # Extract Audiofile
                    video = VideoFileClip(video_file)
                    audio = video.audio
                    # Write audiofiles
                    audio.write_audiofile(audio_file)
                    
                    # Add subject code to Excel after conversion
                    add_subject_to_excel(subject_code, xlsx_file_path)
                except Exception as e:
                    logging.error(f"Error extracting audio from video: {e}")
                    raise

            # Check if the filtered audio file already exists
            if not os.path.exists(filtered_audio_file):
                try:
                    audio = AudioSegment.from_wav(audio_file)

                    # Reduce to one channel if necessary
                    audio = audio.set_channels(1)
                    audio.export(filtered_audio_file, format="wav")
                except Exception as e:
                    logging.error(f"Error processing audio file: {e}")
                    raise


print(f"Processing completed. All subjects saved to {audio_path}")